"""Generate the operator-facing station-table import template."""

from pathlib import Path

from openpyxl import Workbook
from openpyxl.comments import Comment
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side


def create_template(output: Path) -> None:
    """Write a styled, import-ready workbook to output."""
    output.parent.mkdir(parents=True, exist_ok=True)
    workbook = Workbook()
    worksheet = workbook.active
    if worksheet is None:
        raise RuntimeError("New workbook did not create a default worksheet")
    worksheet.title = "Worksheet"
    worksheet.sheet_properties.tabColor = "16A34A"

    headers = ("站位编码", "物料编码")
    worksheet.append(headers)
    worksheet.append(("F-01", "013000081"))

    dark_green = PatternFill("solid", fgColor="166534")
    pale_green = PatternFill("solid", fgColor="F0FDF4")
    white_font = Font(name="Microsoft YaHei", color="FFFFFF", bold=True)
    body_font = Font(name="Microsoft YaHei", color="1F2937")
    border = Border(
        left=Side(style="thin", color="D1D5DB"),
        right=Side(style="thin", color="D1D5DB"),
        top=Side(style="thin", color="D1D5DB"),
        bottom=Side(style="thin", color="D1D5DB"),
    )
    for cell in worksheet[1]:
        cell.fill = dark_green
        cell.font = white_font
        cell.alignment = Alignment(horizontal="center", vertical="center")
        cell.border = border
    for cell in worksheet[2]:
        cell.fill = pale_green
        cell.font = body_font
        cell.alignment = Alignment(horizontal="left", vertical="center")
        cell.border = border
        cell.number_format = "@"
        cell.comment = Comment(
            "Example value - replace or delete this row before import.", "SMT Guard"
        )

    worksheet.row_dimensions[1].height = 28
    worksheet.row_dimensions[2].height = 24
    worksheet.column_dimensions["A"].width = 20
    worksheet.column_dimensions["B"].width = 24
    worksheet.freeze_panes = "A2"
    worksheet.auto_filter.ref = "A1:B2"
    worksheet.print_title_rows = "1:1"
    worksheet.page_setup.orientation = "landscape"
    worksheet.page_setup.fitToWidth = 1
    worksheet.page_setup.fitToHeight = 0
    worksheet.sheet_view.showGridLines = False

    instructions = workbook.create_sheet("填写说明")
    instructions.sheet_properties.tabColor = "64748B"
    instructions.merge_cells("A1:D1")
    instructions["A1"] = "SMT 站位表导入模板 - 填写说明"
    instructions["A1"].fill = dark_green
    instructions["A1"].font = Font(name="Microsoft YaHei", color="FFFFFF", bold=True, size=16)
    instructions["A1"].alignment = Alignment(horizontal="center", vertical="center")
    instructions.row_dimensions[1].height = 34
    notes = (
        ("字段", "是否必填", "示例", "说明"),
        ("站位编码", "是", "F-01", "必须全局唯一，并已在“设备与站位”页面创建和启用"),
        ("物料编码", "是", "013000081", "按文本填写并保留前导零；不要求预先导入 BOM"),
        ("设备编码", "否", "SMT-01", "仅兼容旧表；如填写，必须与该站位的所属设备一致"),
    )
    for row in notes:
        instructions.append(row)
    for cell in instructions[2]:
        cell.fill = PatternFill("solid", fgColor="DCFCE7")
        cell.font = Font(name="Microsoft YaHei", color="166534", bold=True)
        cell.alignment = Alignment(horizontal="center", vertical="center")
        cell.border = border
    for row in instructions.iter_rows(min_row=3, max_row=5, min_col=1, max_col=4):
        for cell in row:
            cell.font = body_font
            cell.alignment = Alignment(vertical="top", wrap_text=True)
            cell.border = border
    instructions.column_dimensions["A"].width = 18
    instructions.column_dimensions["B"].width = 14
    instructions.column_dimensions["C"].width = 20
    instructions.column_dimensions["D"].width = 58
    instructions.freeze_panes = "A3"
    instructions.sheet_view.showGridLines = False
    instructions.page_setup.orientation = "landscape"
    instructions.page_setup.fitToWidth = 1
    instructions.page_setup.fitToHeight = 1

    workbook.save(output)
    workbook.close()


if __name__ == "__main__":
    create_template(Path(__file__).resolve().parents[1] / "templates" / "站位表导入模板.xlsx")
