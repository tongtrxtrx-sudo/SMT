"""Import purchase receipts and maintain reusable Code 128 material-label files."""

# ReportLab does not publish complete type information for its barcode and font APIs.
# pyright: reportUnknownMemberType=false, reportUnknownVariableType=false, reportUnknownArgumentType=false

from __future__ import annotations

import hashlib
import json
import os
import re
import shutil
import tempfile
from collections.abc import Sequence
from dataclasses import dataclass
from datetime import date, datetime
from decimal import Decimal, InvalidOperation
from pathlib import Path
from typing import Final, Literal, cast

from openpyxl import load_workbook
from openpyxl.cell.cell import Cell, MergedCell
from openpyxl.cell.read_only import ReadOnlyCell
from openpyxl.worksheet._read_only import ReadOnlyWorksheet
from openpyxl.worksheet.worksheet import Worksheet
from reportlab.graphics.barcode.code128 import Code128
from reportlab.lib.colors import HexColor, white
from reportlab.lib.units import mm
from reportlab.pdfbase import pdfmetrics
from reportlab.pdfbase.cidfonts import UnicodeCIDFont
from reportlab.pdfbase.ttfonts import TTFError, TTFont
from reportlab.pdfgen.canvas import Canvas

from smt_guard.xlsx_reader import WorkbookReadError, validate_workbook_archive

ReceiptCell = Cell | ReadOnlyCell | MergedCell


class ReceiptImportError(ValueError):
    """Raised when a purchase-receipt workbook is not usable."""


class ReceiptLabelExportError(ValueError):
    """Raised when printable labels cannot be generated safely."""


@dataclass(frozen=True)
class ReceiptItem:
    """One material line imported from a purchase receipt."""

    row_number: int
    material_code: str
    material_name: str
    specification: str
    quantity: str
    unit: str
    category: str
    supplier: str


@dataclass(frozen=True)
class ReceiptDocument:
    """Purchase-receipt metadata and material lines."""

    receipt_number: str
    warehouse: str
    receipt_date: str
    receiver: str
    process_number: str
    source_path: Path
    sheet_name: str
    items: tuple[ReceiptItem, ...]


@dataclass(frozen=True)
class LabelRequest:
    """One selected material that needs a reusable label file."""

    item: ReceiptItem


@dataclass(frozen=True)
class MaterialLabelFormat:
    """One selectable label layout exposed by the material-label tool."""

    format_id: str
    display_name: str
    page_width_mm: int
    page_height_mm: int
    barcode_type: str


@dataclass(frozen=True)
class MaterialLabelFileResult:
    """Library and current-print paths for one selected material."""

    item: ReceiptItem
    library_path: Path
    current_print_path: Path
    reused: bool

    @property
    def output_path(self) -> Path:
        """Return the file users should print now."""
        return self.current_print_path


@dataclass(frozen=True)
class LabelWorkspacePaths:
    """Managed folders below one user-selected label workspace."""

    root: Path
    library_directory: Path
    current_print_directory: Path


@dataclass(frozen=True)
class LabelExportResult:
    """Summary of material-label files generated or reused in one operation."""

    workspace_root: Path
    library_directory: Path
    current_print_directory: Path
    files: tuple[MaterialLabelFileResult, ...]
    label_format_id: str = "60x40-code128-v3"

    @property
    def output_directory(self) -> Path:
        """Return the folder containing only the labels ready to print now."""
        return self.current_print_directory

    @property
    def label_count(self) -> int:
        """Return the number of one-page material-label files."""
        return len(self.files)

    @property
    def generated_count(self) -> int:
        """Return how many files were created or refreshed."""
        return sum(not item.reused for item in self.files)

    @property
    def reused_count(self) -> int:
        """Return how many existing files were reused without modification."""
        return sum(item.reused for item in self.files)


class ReceiptLabelWorkspaceSettings:
    """Persist the selected label workspace outside the SMT Guard database."""

    SETTINGS_VERSION: Final[int] = 2
    LEGACY_SETTINGS_VERSION: Final[int] = 1

    def __init__(
        self,
        settings_path: Path | None = None,
        default_workspace_root: Path | None = None,
    ) -> None:
        local_app_data = Path(
            os.environ.get(
                "LOCALAPPDATA",
                str(Path.home() / "AppData" / "Local"),
            )
        )
        self.settings_path = (
            settings_path
            if settings_path is not None
            else local_app_data / "SMTReceiptLabels" / "settings.json"
        ).expanduser()
        self.default_workspace_root = (
            default_workspace_root
            if default_workspace_root is not None
            else Path.home() / "Desktop" / "SMT物料标签"
        ).expanduser()

    def load_workspace_root(self) -> Path:
        """Return a saved root, migrating the former Documents default."""
        try:
            payload = json.loads(self.settings_path.read_text(encoding="utf-8"))
        except (OSError, UnicodeError, json.JSONDecodeError):
            return self.default_workspace_root.resolve()
        if not isinstance(payload, dict):
            return self.default_workspace_root.resolve()
        settings_version = payload.get("settings_version")
        if settings_version not in (
            self.LEGACY_SETTINGS_VERSION,
            self.SETTINGS_VERSION,
        ):
            return self.default_workspace_root.resolve()
        root = payload.get("workspace_root")
        if not isinstance(root, str) or not root.strip():
            return self.default_workspace_root.resolve()
        saved_root = Path(root).expanduser().resolve()
        if (
            settings_version == self.LEGACY_SETTINGS_VERSION
            and saved_root.name == "SMT物料标签"
            and saved_root.parent.name.casefold() == "documents"
        ):
            return self.default_workspace_root.resolve()
        return saved_root

    def save_workspace_root(self, workspace_root: Path) -> None:
        """Atomically save the root used for the next startup cleanup."""
        target = self.settings_path.resolve()
        try:
            target.parent.mkdir(parents=True, exist_ok=True)
            descriptor, temporary_name = tempfile.mkstemp(
                prefix=".settings-",
                suffix=".tmp",
                dir=target.parent,
            )
            os.close(descriptor)
            temporary = Path(temporary_name)
            try:
                temporary.write_text(
                    json.dumps(
                        {
                            "settings_version": self.SETTINGS_VERSION,
                            "workspace_root": str(workspace_root.expanduser().resolve()),
                        },
                        ensure_ascii=False,
                        indent=2,
                    )
                    + "\n",
                    encoding="utf-8",
                )
                temporary.replace(target)
            finally:
                temporary.unlink(missing_ok=True)
        except OSError as error:
            raise ReceiptLabelExportError(f"无法保存标签工作目录设置：{error}") from error


class ReceiptWorkbookImporter:
    """Read a purchase-receipt export by its Chinese column names."""

    REQUIRED_HEADERS: Final[frozenset[str]] = frozenset(
        {"采购入库单号", "商品名", "商品编号", "入库数量"}
    )
    OPTIONAL_HEADERS: Final[tuple[str, ...]] = (
        "仓库",
        "入库日期",
        "收货人",
        "商品规格",
        "商品类别",
        "商品供应商",
        "商品单位",
        "加工单号",
    )
    HEADER_SEARCH_ROWS: Final[int] = 20

    def import_file(self, path: Path) -> ReceiptDocument:
        """Import one workbook without modifying it."""
        source_path = path.expanduser().resolve()
        try:
            validate_workbook_archive(source_path)
            workbook = load_workbook(source_path, read_only=True, data_only=True)
        except WorkbookReadError as error:
            raise ReceiptImportError(str(error)) from error
        except (OSError, ValueError) as error:
            raise ReceiptImportError(f"无法打开入库单：{error}") from error

        try:
            candidates: list[tuple[str, int, dict[str, int]]] = []
            for sheet_name in workbook.sheetnames:
                worksheet = workbook[sheet_name]
                match = self._find_header(worksheet)
                if match is not None:
                    header_row, columns = match
                    candidates.append((sheet_name, header_row, columns))
            if not candidates:
                required = "、".join(sorted(self.REQUIRED_HEADERS))
                raise ReceiptImportError(f"未找到入库明细表，必须包含列：{required}")
            if len(candidates) > 1:
                names = "、".join(candidate[0] for candidate in candidates)
                raise ReceiptImportError(f"检测到多个入库明细工作表，请只保留一个：{names}")

            sheet_name, header_row, columns = candidates[0]
            worksheet = workbook[sheet_name]
            rows = list(
                worksheet.iter_rows(
                    min_row=header_row + 1,
                    max_row=worksheet.max_row,
                    max_col=worksheet.max_column,
                )
            )
            if not rows:
                raise ReceiptImportError("入库单没有物料明细")

            receipt_numbers = self._column_values(rows, columns["采购入库单号"])
            distinct_receipts = tuple(dict.fromkeys(receipt_numbers))
            if len(distinct_receipts) > 1:
                raise ReceiptImportError("一个文件中包含多个采购入库单号，请分别导出后再打印")

            items: list[ReceiptItem] = []
            for row_offset, row in enumerate(rows, start=header_row + 1):
                code = self._row_text(row, columns["商品编号"])
                name = self._row_text(row, columns["商品名"])
                if not code and not name:
                    continue
                if not code:
                    raise ReceiptImportError(f"第 {row_offset} 行商品编号为空")
                if not name:
                    raise ReceiptImportError(f"第 {row_offset} 行商品名为空")
                items.append(
                    ReceiptItem(
                        row_number=row_offset,
                        material_code=code,
                        material_name=name,
                        specification=self._optional_row_text(row, columns, "商品规格"),
                        quantity=self._row_text(row, columns["入库数量"]),
                        unit=self._optional_row_text(row, columns, "商品单位"),
                        category=self._optional_row_text(row, columns, "商品类别"),
                        supplier=self._optional_row_text(row, columns, "商品供应商"),
                    )
                )
            if not items:
                raise ReceiptImportError("入库单没有可打印的物料明细")

            return ReceiptDocument(
                receipt_number=distinct_receipts[0] if distinct_receipts else "",
                warehouse=self._first_column_value(rows, columns, "仓库"),
                receipt_date=self._first_column_value(rows, columns, "入库日期"),
                receiver=self._first_column_value(rows, columns, "收货人"),
                process_number=self._first_column_value(rows, columns, "加工单号"),
                source_path=source_path,
                sheet_name=sheet_name,
                items=tuple(items),
            )
        finally:
            workbook.close()

    def _find_header(
        self,
        worksheet: Worksheet | ReadOnlyWorksheet,
    ) -> tuple[int, dict[str, int]] | None:
        for row_number, row in enumerate(
            worksheet.iter_rows(min_row=1, max_row=self.HEADER_SEARCH_ROWS), start=1
        ):
            columns: dict[str, int] = {}
            for index, cell in enumerate(row):
                header = self._cell_text(cell)
                if header and header not in columns:
                    columns[header] = index
            if self.REQUIRED_HEADERS.issubset(columns):
                return row_number, columns
        return None

    @classmethod
    def _column_values(
        cls,
        rows: Sequence[Sequence[ReceiptCell]],
        index: int,
    ) -> list[str]:
        return [value for row in rows if (value := cls._row_text(row, index))]

    @classmethod
    def _first_column_value(
        cls,
        rows: Sequence[Sequence[ReceiptCell]],
        columns: dict[str, int],
        header: str,
    ) -> str:
        index = columns.get(header)
        if index is None:
            return ""
        return next((value for row in rows if (value := cls._row_text(row, index))), "")

    @classmethod
    def _optional_row_text(
        cls,
        row: Sequence[ReceiptCell],
        columns: dict[str, int],
        header: str,
    ) -> str:
        index = columns.get(header)
        return "" if index is None else cls._row_text(row, index)

    @classmethod
    def _row_text(cls, row: Sequence[ReceiptCell], index: int) -> str:
        if index >= len(row):
            return ""
        return cls._cell_text(row[index])

    @staticmethod
    def _cell_text(cell: ReceiptCell) -> str:
        value = cell.value
        if value is None:
            return ""
        if isinstance(value, datetime):
            return value.strftime("%Y-%m-%d %H:%M:%S")
        if isinstance(value, date):
            return value.isoformat()
        if isinstance(value, str):
            return value.strip()
        if isinstance(value, bool):
            return "是" if value else "否"
        if isinstance(value, int | float | Decimal):
            number_format = str(cell.number_format or "").split(";", maxsplit=1)[0]
            if re.fullmatch(r"0+", number_format):
                try:
                    integer = int(Decimal(str(value)))
                except (InvalidOperation, ValueError):
                    pass
                else:
                    return f"{integer:0{len(number_format)}d}"
            try:
                decimal = Decimal(str(value))
            except InvalidOperation:
                return str(value).strip()
            if decimal == decimal.to_integral():
                return str(int(decimal))
            return format(decimal.normalize(), "f")
        return str(value).strip()


class ReceiptLabelPdfExporter:
    """Maintain reusable labels and prepare a clean current-print folder."""

    FONT_NAME: Final[str] = "SMTReceiptChinese"
    FALLBACK_FONT_NAME: Final[str] = "STSong-Light"
    PAGE_WIDTH: Final[float] = 60 * mm
    PAGE_HEIGHT: Final[float] = 40 * mm
    PAGE_SIZE: Final[tuple[float, float]] = (PAGE_WIDTH, PAGE_HEIGHT)
    MATERIAL_CODE_BASELINE: Final[float] = 12.7 * mm
    SPECIFICATION_FIRST_BASELINE: Final[float] = 8.8 * mm
    CACHE_FILENAME: Final[str] = ".smt-material-labels.json"
    CACHE_SCHEMA_VERSION: Final[int] = 1
    # v3 adds clearer vertical spacing below the material code. Bumping the
    # format id invalidates cached PDFs so the updated layout is regenerated.
    DEFAULT_LABEL_FORMAT_ID: Final[str] = "60x40-code128-v3"
    LABEL_FORMAT_VERSION: Final[str] = DEFAULT_LABEL_FORMAT_ID
    LABEL_FORMATS: Final[tuple[MaterialLabelFormat, ...]] = (
        MaterialLabelFormat(
            format_id=DEFAULT_LABEL_FORMAT_ID,
            display_name="60 × 40 mm · Code 128 标准物料标签",
            page_width_mm=60,
            page_height_mm=40,
            barcode_type="Code 128",
        ),
    )
    LIBRARY_DIRECTORY_NAME: Final[str] = "标签库"
    CURRENT_PRINT_DIRECTORY_NAME: Final[str] = "当前打印"
    WORKSPACE_MARKER: Final[str] = ".smt-label-workspace"
    LEGACY_CURRENT_PRINT_MARKER: Final[str] = ".smt-current-print"

    def __init__(self, font_path: Path | None = None) -> None:
        self._font_path = font_path
        self._font_name = self.FONT_NAME

    def export(
        self,
        requests: tuple[LabelRequest, ...],
        workspace_root: Path,
        *,
        label_format_id: str | None = None,
    ) -> LabelExportResult:
        """Update the library, then copy only this selection to current print."""
        label_format = self._resolve_format(label_format_id)
        items = self._unique_items(requests)
        paths = self.prepare_workspace(workspace_root, clear_current_print=False)

        self._font_name = self._register_font()
        manifest = self._load_manifest(
            paths.library_directory,
            label_format.format_id,
        )
        cache_entries = cast(dict[str, dict[str, str]], manifest["materials"])
        library_results: list[tuple[ReceiptItem, Path, bool]] = []
        for item in items:
            filename = self._filename_for(item.material_code)
            target = paths.library_directory / filename
            fingerprint = self._fingerprint(item, label_format.format_id)
            entry = cache_entries.get(item.material_code)
            reusable = self._entry_is_reusable(entry, target, filename, fingerprint)
            if not reusable:
                self._write_label_atomically(item, target, fingerprint)
            cache_entries[item.material_code] = {
                "filename": filename,
                "fingerprint": fingerprint,
                "material_name": item.material_name,
                "specification": item.specification,
            }
            library_results.append((item, target, reusable))

        self._write_manifest(paths.library_directory, manifest)
        self._clear_current_print(paths)
        results: list[MaterialLabelFileResult] = []
        for item, library_path, reusable in library_results:
            current_print_path = paths.current_print_directory / library_path.name
            self._copy_label_atomically(library_path, current_print_path)
            results.append(
                MaterialLabelFileResult(
                    item,
                    library_path,
                    current_print_path,
                    reusable,
                )
            )
        return LabelExportResult(
            paths.root,
            paths.library_directory,
            paths.current_print_directory,
            tuple(results),
            label_format.format_id,
        )

    @classmethod
    def available_formats(cls) -> tuple[MaterialLabelFormat, ...]:
        """Return selectable formats in their intended UI order."""
        return cls.LABEL_FORMATS

    @classmethod
    def _resolve_format(cls, label_format_id: str | None) -> MaterialLabelFormat:
        selected_id = label_format_id or cls.DEFAULT_LABEL_FORMAT_ID
        for label_format in cls.LABEL_FORMATS:
            if label_format.format_id == selected_id:
                return label_format
        raise ReceiptLabelExportError(f"不支持的标签格式：{selected_id}")

    def prepare_workspace(
        self,
        workspace_root: Path,
        *,
        clear_current_print: bool,
    ) -> LabelWorkspacePaths:
        """Create protected managed folders and optionally clear current files."""
        paths = self.workspace_paths(workspace_root)
        if paths.root.exists() and not paths.root.is_dir():
            raise ReceiptLabelExportError("标签工作目录必须是文件夹")
        try:
            paths.library_directory.mkdir(parents=True, exist_ok=True)
            paths.current_print_directory.mkdir(parents=True, exist_ok=True)
        except OSError as error:
            raise ReceiptLabelExportError(f"无法创建标签工作目录：{error}") from error

        marker = paths.root / self.WORKSPACE_MARKER
        legacy_marker = (
            paths.current_print_directory / self.LEGACY_CURRENT_PRINT_MARKER
        )
        if not marker.is_file():
            try:
                existing = tuple(paths.current_print_directory.iterdir())
            except OSError as error:
                raise ReceiptLabelExportError(f"无法检查当前打印文件夹：{error}") from error
            if existing and not legacy_marker.is_file():
                raise ReceiptLabelExportError(
                    "当前打印文件夹中已有非本软件管理的内容；为防止误删，请先移走这些文件"
                )
            try:
                marker.write_text(
                    "此工作目录由 SMT 物料标签库管理。\n",
                    encoding="utf-8",
                )
                self._hide_windows_marker(marker)
            except OSError as error:
                raise ReceiptLabelExportError(f"无法保护当前打印文件夹：{error}") from error
        if legacy_marker.is_file():
            try:
                legacy_marker.unlink()
            except OSError as error:
                raise ReceiptLabelExportError(f"无法升级当前打印文件夹：{error}") from error
        if clear_current_print:
            self._clear_current_print(paths)
        return paths

    @staticmethod
    def _hide_windows_marker(marker: Path) -> None:
        """Hide the safety marker in ordinary Windows Explorer views."""
        if os.name != "nt":
            return
        try:
            import ctypes

            kernel32 = ctypes.WinDLL("kernel32", use_last_error=True)
            get_attributes = kernel32.GetFileAttributesW
            set_attributes = kernel32.SetFileAttributesW
            attributes = int(get_attributes(str(marker)))
            if attributes != 0xFFFFFFFF:
                set_attributes(str(marker), attributes | 0x2)
        except (AttributeError, OSError, ValueError):
            return

    @classmethod
    def workspace_paths(cls, workspace_root: Path) -> LabelWorkspacePaths:
        """Resolve the two fixed managed subfolders without touching the disk."""
        root = workspace_root.expanduser().resolve()
        return LabelWorkspacePaths(
            root,
            root / cls.LIBRARY_DIRECTORY_NAME,
            root / cls.CURRENT_PRINT_DIRECTORY_NAME,
        )

    def status(
        self,
        item: ReceiptItem,
        workspace_root: Path,
        *,
        label_format_id: str | None = None,
    ) -> Literal["new", "reusable", "update"]:
        """Inspect whether a material file is new, reusable, or needs refreshing."""
        label_format = self._resolve_format(label_format_id)
        directory = self.workspace_paths(workspace_root).library_directory
        filename = self._filename_for(item.material_code)
        target = directory / filename
        manifest = self._load_manifest(directory, label_format.format_id)
        entries = cast(dict[str, dict[str, str]], manifest["materials"])
        entry = entries.get(item.material_code)
        if self._entry_is_reusable(
            entry,
            target,
            filename,
            self._fingerprint(item, label_format.format_id),
        ):
            return "reusable"
        return "update" if target.exists() else "new"

    @classmethod
    def _clear_current_print(cls, paths: LabelWorkspacePaths) -> None:
        """Remove current-print contents only after verifying the workspace marker."""
        marker = paths.root / cls.WORKSPACE_MARKER
        if not marker.is_file():
            raise ReceiptLabelExportError(
                "标签工作目录缺少安全标记，已停止自动清理以避免误删文件"
            )
        try:
            entries = tuple(paths.current_print_directory.iterdir())
            for entry in entries:
                if entry.is_file() or entry.is_symlink():
                    entry.unlink()
                elif entry.is_dir():
                    shutil.rmtree(entry)
        except OSError as error:
            raise ReceiptLabelExportError(f"无法清空当前打印文件夹：{error}") from error

    @staticmethod
    def _copy_label_atomically(source: Path, target: Path) -> None:
        descriptor, temporary_name = tempfile.mkstemp(
            prefix=f".{target.stem}-",
            suffix=".tmp.pdf",
            dir=target.parent,
        )
        os.close(descriptor)
        temporary = Path(temporary_name)
        try:
            shutil.copy2(source, temporary)
            temporary.replace(target)
        except OSError as error:
            raise ReceiptLabelExportError(
                f"无法把 {source.name} 放入当前打印文件夹：{error}"
            ) from error
        finally:
            temporary.unlink(missing_ok=True)

    @staticmethod
    def _unique_items(requests: tuple[LabelRequest, ...]) -> tuple[ReceiptItem, ...]:
        if not requests:
            raise ReceiptLabelExportError("请至少选择一种需要生成标签的物料")
        unique: dict[str, ReceiptItem] = {}
        for request in requests:
            item = request.item
            code = item.material_code.strip()
            if not code:
                raise ReceiptLabelExportError("商品编号不能为空")
            existing = unique.get(code)
            if existing is not None and (
                existing.material_name != item.material_name
                or existing.specification != item.specification
            ):
                raise ReceiptLabelExportError(
                    f"商品编号 {code} 对应了不同的名称或规格，请先核对入库单"
                )
            unique.setdefault(code, item)
        return tuple(unique.values())

    @classmethod
    def _filename_for(cls, material_code: str) -> str:
        code = material_code.strip()
        safe_code = re.sub(r'[<>:"/\\|?*\x00-\x1f]', "_", code).strip(" .")
        if not safe_code:
            safe_code = hashlib.sha256(code.encode("utf-8")).hexdigest()[:12]
        elif safe_code != code:
            suffix = hashlib.sha256(code.encode("utf-8")).hexdigest()[:8]
            safe_code = f"{safe_code}-{suffix}"
        return f"物料-{safe_code}.pdf"

    @classmethod
    def _fingerprint(cls, item: ReceiptItem, label_format_id: str) -> str:
        payload = json.dumps(
            {
                "format": label_format_id,
                "material_code": item.material_code,
                "material_name": item.material_name,
                "specification": item.specification,
            },
            ensure_ascii=False,
            sort_keys=True,
            separators=(",", ":"),
        )
        return hashlib.sha256(payload.encode("utf-8")).hexdigest()

    @classmethod
    def _empty_manifest(cls, label_format_id: str) -> dict[str, object]:
        return {
            "schema_version": cls.CACHE_SCHEMA_VERSION,
            "label_format": label_format_id,
            "materials": {},
        }

    @classmethod
    def _load_manifest(
        cls,
        directory: Path,
        label_format_id: str,
    ) -> dict[str, object]:
        path = directory / cls.CACHE_FILENAME
        if not path.is_file():
            return cls._empty_manifest(label_format_id)
        try:
            payload = json.loads(path.read_text(encoding="utf-8"))
        except (OSError, UnicodeError, json.JSONDecodeError):
            return cls._empty_manifest(label_format_id)
        if not isinstance(payload, dict):
            return cls._empty_manifest(label_format_id)
        if payload.get("schema_version") != cls.CACHE_SCHEMA_VERSION:
            return cls._empty_manifest(label_format_id)
        if payload.get("label_format") != label_format_id:
            return cls._empty_manifest(label_format_id)
        raw_materials = payload.get("materials")
        if not isinstance(raw_materials, dict):
            return cls._empty_manifest(label_format_id)

        materials: dict[str, dict[str, str]] = {}
        for code, raw_entry in raw_materials.items():
            if not isinstance(code, str) or not isinstance(raw_entry, dict):
                continue
            entry = {
                key: value
                for key, value in raw_entry.items()
                if isinstance(key, str) and isinstance(value, str)
            }
            materials[code] = entry
        return {
            "schema_version": cls.CACHE_SCHEMA_VERSION,
            "label_format": label_format_id,
            "materials": materials,
        }

    @staticmethod
    def _entry_is_reusable(
        entry: object,
        target: Path,
        filename: str,
        fingerprint: str,
    ) -> bool:
        return (
            isinstance(entry, dict)
            and entry.get("filename") == filename
            and entry.get("fingerprint") == fingerprint
            and target.is_file()
            and target.stat().st_size > 0
        )

    def _write_label_atomically(
        self,
        item: ReceiptItem,
        target: Path,
        fingerprint: str,
    ) -> None:
        descriptor, temporary_name = tempfile.mkstemp(
            prefix=f".{target.stem}-",
            suffix=".tmp.pdf",
            dir=target.parent,
        )
        os.close(descriptor)
        temporary = Path(temporary_name)
        try:
            self._write_label(item, temporary, fingerprint)
            temporary.replace(target)
        except (OSError, ValueError) as error:
            raise ReceiptLabelExportError(
                f"无法生成物料 {item.material_code} 的标签：{error}"
            ) from error
        finally:
            temporary.unlink(missing_ok=True)

    def _write_label(self, item: ReceiptItem, target: Path, fingerprint: str) -> None:
        canvas = Canvas(str(target), pagesize=self.PAGE_SIZE, pageCompression=1)
        canvas.setTitle(f"物料标签 {item.material_code}")
        canvas.setAuthor("SMT 物料标签工具")
        canvas.setSubject(f"60 × 40 mm reusable material label; {fingerprint}")
        self._draw_label(canvas, item)
        canvas.save()

    def _draw_label(self, canvas: Canvas, item: ReceiptItem) -> None:
        padding = 2.5 * mm
        content_width = self.PAGE_WIDTH - 2 * padding

        canvas.setFillColor(white)
        canvas.rect(0, 0, self.PAGE_WIDTH, self.PAGE_HEIGHT, fill=1, stroke=0)
        canvas.setStrokeColor(HexColor("#AAB8C8"))
        canvas.setLineWidth(0.45)
        canvas.roundRect(
            1 * mm,
            1 * mm,
            self.PAGE_WIDTH - 2 * mm,
            self.PAGE_HEIGHT - 2 * mm,
            1.5 * mm,
            fill=0,
            stroke=1,
        )

        canvas.setFillColor(HexColor("#0F2747"))
        canvas.setFont(self._font_name, 9.5)
        canvas.drawString(
            padding,
            34.4 * mm,
            self._fit_text(item.material_name, self._font_name, 9.5, content_width),
        )

        try:
            barcode = Code128(
                item.material_code,
                barWidth=0.38 * mm,
                barHeight=13 * mm,
                humanReadable=False,
                quiet=True,
            )
        except (IndexError, TypeError, ValueError) as error:
            raise ReceiptLabelExportError(
                f"商品编号 {item.material_code} 无法编码为 Code 128"
            ) from error
        scale = min(1.0, content_width / barcode.width)
        if scale < 0.65:
            raise ReceiptLabelExportError(
                f"商品编号 {item.material_code} 太长，无法生成可可靠扫描的 Code 128 标签"
            )
        canvas.saveState()
        barcode_x = (self.PAGE_WIDTH - barcode.width * scale) / 2
        canvas.translate(barcode_x, 17 * mm)
        canvas.scale(scale, 1)
        barcode.drawOn(canvas, 0, 0)
        canvas.restoreState()

        canvas.setFillColor(HexColor("#111827"))
        canvas.setFont("Helvetica-Bold", 10)
        canvas.drawCentredString(
            self.PAGE_WIDTH / 2,
            self.MATERIAL_CODE_BASELINE,
            item.material_code,
        )

        canvas.setFillColor(HexColor("#334155"))
        specification = item.specification.strip() or "-"
        specification_lines, specification_font_size = self._fit_text_lines(
            f"规格：{specification}",
            self._font_name,
            preferred_font_size=6.8,
            minimum_font_size=4.8,
            max_width=content_width,
            max_lines=4,
        )
        canvas.setFont(self._font_name, specification_font_size)
        line_height = specification_font_size + 0.7
        first_baseline = self.SPECIFICATION_FIRST_BASELINE
        for line_number, line in enumerate(specification_lines):
            canvas.drawString(
                padding,
                first_baseline - line_number * line_height,
                line,
            )

    @staticmethod
    def _fit_text(value: str, font_name: str, font_size: float, max_width: float) -> str:
        text = value.strip()
        if pdfmetrics.stringWidth(text, font_name, font_size) <= max_width:
            return text
        ellipsis = "…"
        while (
            text and pdfmetrics.stringWidth(f"{text}{ellipsis}", font_name, font_size) > max_width
        ):
            text = text[:-1]
        return f"{text}{ellipsis}" if text else ellipsis

    @classmethod
    def _fit_text_lines(
        cls,
        value: str,
        font_name: str,
        *,
        preferred_font_size: float,
        minimum_font_size: float,
        max_width: float,
        max_lines: int,
    ) -> tuple[tuple[str, ...], float]:
        """Wrap all text, reducing the font only when the line limit requires it."""
        font_size = preferred_font_size
        while font_size >= minimum_font_size - 0.001:
            lines = cls._wrap_text(value, font_name, font_size, max_width)
            if len(lines) <= max_lines:
                return lines, font_size
            font_size = round(font_size - 0.2, 1)
        raise ReceiptLabelExportError(
            f"规格内容过长，无法在当前标签格式中完整显示（最多 {max_lines} 行）"
        )

    @staticmethod
    def _wrap_text(
        value: str,
        font_name: str,
        font_size: float,
        max_width: float,
    ) -> tuple[str, ...]:
        text = value.strip()
        if not text:
            return ("",)
        lines: list[str] = []
        current = ""
        for word in re.findall(r"\S+", text):
            candidate = f"{current} {word}" if current else word
            if pdfmetrics.stringWidth(candidate, font_name, font_size) <= max_width:
                current = candidate
                continue

            if current:
                lines.append(current)
                current = ""

            # Preserve complete model numbers and other whitespace-delimited
            # tokens whenever possible. Only a single overlong token (usually
            # continuous Chinese text) falls back to character wrapping.
            if pdfmetrics.stringWidth(word, font_name, font_size) <= max_width:
                current = word
                continue
            for character in word:
                candidate = f"{current}{character}"
                if current and pdfmetrics.stringWidth(
                    candidate,
                    font_name,
                    font_size,
                ) > max_width:
                    lines.append(current)
                    current = character
                else:
                    current = candidate
        if current:
            lines.append(current)
        return tuple(lines)

    @classmethod
    def _write_manifest(cls, directory: Path, manifest: dict[str, object]) -> None:
        target = directory / cls.CACHE_FILENAME
        descriptor, temporary_name = tempfile.mkstemp(
            prefix=".smt-material-labels-",
            suffix=".tmp",
            dir=directory,
        )
        os.close(descriptor)
        temporary = Path(temporary_name)
        try:
            temporary.write_text(
                json.dumps(manifest, ensure_ascii=False, indent=2) + "\n",
                encoding="utf-8",
            )
            temporary.replace(target)
        except OSError as error:
            raise ReceiptLabelExportError(f"无法保存标签复用记录：{error}") from error
        finally:
            temporary.unlink(missing_ok=True)

    def _register_font(self) -> str:
        """Embed a Windows Chinese font when possible for reliable printing."""
        if self.FONT_NAME in pdfmetrics.getRegisteredFontNames():
            return self.FONT_NAME
        candidates = (
            (self._font_path, 0),
            (Path(os.environ.get("WINDIR", r"C:\Windows")) / "Fonts" / "msyh.ttc", 0),
            (Path(os.environ.get("WINDIR", r"C:\Windows")) / "Fonts" / "simsun.ttc", 0),
        )
        for path, subfont_index in candidates:
            if path is None or not path.is_file():
                continue
            try:
                pdfmetrics.registerFont(
                    TTFont(
                        self.FONT_NAME,
                        str(path),
                        subfontIndex=subfont_index,
                    )
                )
            except (OSError, TTFError):
                continue
            return self.FONT_NAME
        if self.FALLBACK_FONT_NAME not in pdfmetrics.getRegisteredFontNames():
            pdfmetrics.registerFont(UnicodeCIDFont(self.FALLBACK_FONT_NAME))
        return self.FALLBACK_FONT_NAME
