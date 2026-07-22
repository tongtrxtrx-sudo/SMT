"""OpenPyXL adapter for reading import worksheets safely."""

from collections.abc import Iterable, Sequence
from dataclasses import dataclass
from pathlib import Path
from typing import Protocol, cast
from zipfile import BadZipFile, ZipFile


class WorkbookReadError(ValueError):
    """Raised when a workbook cannot be mapped to unambiguous tabular rows."""


class AmbiguousWorksheetError(WorkbookReadError):
    """Raised when automatic worksheet selection requires an operator choice."""

    def __init__(self, sheet_names: Sequence[str]) -> None:
        self.sheet_names = tuple(sheet_names)
        available = "、".join(self.sheet_names)
        super().__init__(f"工作簿包含多个工作表，请选择要导入的页签：{available}")


@dataclass(frozen=True)
class WorkbookLimits:
    """Resource limits applied before an XLSX ZIP container is extracted."""

    max_archive_bytes: int = 50 * 1024 * 1024
    max_uncompressed_bytes: int = 200 * 1024 * 1024
    max_compression_ratio: float = 1000


DEFAULT_WORKBOOK_LIMITS = WorkbookLimits()


def validate_workbook_archive(
    path: Path,
    *,
    limits: WorkbookLimits = DEFAULT_WORKBOOK_LIMITS,
) -> None:
    """Reject unsupported or resource-exhausting workbook containers."""
    if path.suffix.lower() != ".xlsx":
        raise WorkbookReadError(f"Workbook must use the .xlsx extension: {path}")
    try:
        archive_size = path.stat().st_size
        if archive_size > limits.max_archive_bytes:
            raise WorkbookReadError(
                f"Workbook archive exceeds {limits.max_archive_bytes} bytes: {path}"
            )
        with ZipFile(path) as archive:
            members = archive.infolist()
            uncompressed_size = sum(member.file_size for member in members)
            if uncompressed_size > limits.max_uncompressed_bytes:
                raise WorkbookReadError(
                    "Workbook uncompressed content exceeds "
                    f"{limits.max_uncompressed_bytes} bytes: {path}"
                )
            for member in members:
                if member.flag_bits & 0x1:
                    raise WorkbookReadError(f"Encrypted workbook entries are not supported: {path}")
                if member.file_size and not member.compress_size:
                    raise WorkbookReadError(
                        f"Workbook contains an invalid compressed entry: {path}"
                    )
                if member.compress_size:
                    ratio = member.file_size / member.compress_size
                    if ratio > limits.max_compression_ratio:
                        raise WorkbookReadError(
                            f"Workbook compression ratio is suspicious: {path}"
                        )
    except WorkbookReadError:
        raise
    except (BadZipFile, OSError) as error:
        raise WorkbookReadError(f"Cannot inspect workbook archive {path}: {error}") from error


class WorksheetLike(Protocol):
    """Subset of an OpenPyXL worksheet used by the adapter."""

    def iter_rows(self, *, values_only: bool) -> Iterable[Sequence[object | None]]:
        """Yield worksheet rows."""
        ...


class WorkbookLike(Protocol):
    """Subset of an OpenPyXL workbook used by the adapter."""

    @property
    def sheetnames(self) -> Sequence[str]:
        """Return worksheet names."""
        ...

    def __getitem__(self, name: str) -> WorksheetLike:
        """Return a worksheet by name."""
        ...

    def close(self) -> None:
        """Release workbook resources."""
        ...


class WorkbookLoader(Protocol):
    """Callable boundary around OpenPyXL's workbook loader."""

    def __call__(
        self,
        path: Path,
        *,
        read_only: bool,
        data_only: bool,
    ) -> WorkbookLike:
        """Open a workbook with the requested safety options."""
        ...


def _load_workbook(path: Path, *, read_only: bool, data_only: bool) -> WorkbookLike:
    validate_workbook_archive(path)
    try:
        from openpyxl import load_workbook
    except ModuleNotFoundError as error:
        raise WorkbookReadError(
            "OpenPyXL is required to read .xlsx files; install the project dependencies first"
        ) from error

    return cast(
        WorkbookLike,
        load_workbook(filename=path, read_only=read_only, data_only=data_only),
    )


class OpenpyxlWorkbookReader:
    """Read a named worksheet into normalized row dictionaries."""

    _ROW_NUMBER = "_row_number"

    def __init__(self, workbook_loader: WorkbookLoader = _load_workbook) -> None:
        self._workbook_loader = workbook_loader

    def read_sheet(self, path: Path, sheet_name: str) -> list[dict[str, object]]:
        workbook = self._open(path)
        try:
            selected_sheet = self._select_sheet(workbook.sheetnames, sheet_name)
            if selected_sheet not in workbook.sheetnames:
                available = ", ".join(workbook.sheetnames) or "none"
                raise WorkbookReadError(
                    f"Worksheet {selected_sheet!r} was not found; available worksheets: {available}"
                )

            rows = iter(workbook[selected_sheet].iter_rows(values_only=True))
            try:
                raw_headers = next(rows)
            except StopIteration as error:
                raise WorkbookReadError(f"Worksheet {selected_sheet!r} is empty") from error

            headers = self._headers(raw_headers)
            return self._map_rows(rows, headers)
        finally:
            workbook.close()

    @staticmethod
    def _select_sheet(sheet_names: Sequence[str], requested: str) -> str:
        normalized = requested.strip()
        if normalized:
            return normalized
        if "Worksheet" in sheet_names:
            return "Worksheet"
        if len(sheet_names) == 1:
            return sheet_names[0]
        if not sheet_names:
            raise WorkbookReadError("工作簿中没有可导入的工作表")
        raise AmbiguousWorksheetError(sheet_names)

    def _open(self, path: Path) -> WorkbookLike:
        try:
            return self._workbook_loader(path, read_only=True, data_only=True)
        except WorkbookReadError:
            raise
        except Exception as error:
            raise WorkbookReadError(f"Cannot open workbook {path}: {error}") from error

    @classmethod
    def _headers(cls, values: Sequence[object | None]) -> list[str]:
        headers = [cls._text(value) for value in values]
        if not headers or any(not header for header in headers):
            raise WorkbookReadError("Worksheet header row contains a blank column name")
        if cls._ROW_NUMBER in headers:
            raise WorkbookReadError(f"Worksheet uses reserved column name {cls._ROW_NUMBER}")
        if len(headers) != len(set(headers)):
            raise WorkbookReadError("Worksheet header row contains duplicate column names")
        return headers

    @classmethod
    def _map_rows(
        cls,
        rows: Iterable[Sequence[object | None]],
        headers: Sequence[str],
    ) -> list[dict[str, object]]:
        mapped_rows: list[dict[str, object]] = []
        for row_number, values in enumerate(rows, start=2):
            normalized = cls._normalize_row(values, len(headers))
            if not any(normalized):
                continue
            row: dict[str, object] = dict(zip(headers, normalized, strict=True))
            row[cls._ROW_NUMBER] = row_number
            mapped_rows.append(row)
        return mapped_rows

    @classmethod
    def _normalize_row(cls, values: Sequence[object | None], width: int) -> list[str]:
        return [cls._text(values[index]) if index < len(values) else "" for index in range(width)]

    @staticmethod
    def _text(value: object | None) -> str:
        return "" if value is None else str(value).strip()
