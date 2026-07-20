import unittest
from pathlib import Path

from openpyxl import load_workbook


class StationTableTemplateTests(unittest.TestCase):
    def setUp(self) -> None:
        self.project_root = Path(__file__).resolve().parents[2]
        self.template = self.project_root / "templates" / "站位表导入模板.xlsx"

    def test_template_has_importable_text_safe_data_sheet(self) -> None:
        workbook = load_workbook(self.template, data_only=False)
        self.addCleanup(workbook.close)

        self.assertIn("Worksheet", workbook.sheetnames)
        worksheet = workbook["Worksheet"]
        self.assertEqual(
            ["站位编码", "物料编码"],
            [worksheet.cell(1, column).value for column in range(1, 3)],
        )
        self.assertEqual("013000081", worksheet["B2"].value)
        self.assertEqual("@", worksheet["B2"].number_format)
        self.assertEqual("A2", worksheet.freeze_panes)
        self.assertEqual("A1:B2", worksheet.auto_filter.ref)

    def test_template_is_styled_documented_and_print_ready(self) -> None:
        workbook = load_workbook(self.template, data_only=False)
        self.addCleanup(workbook.close)

        worksheet = workbook["Worksheet"]
        instructions = workbook["填写说明"]
        self.assertNotEqual("00000000", worksheet["A1"].fill.fgColor.rgb)
        self.assertGreaterEqual(worksheet.column_dimensions["B"].width or 0, 18)
        self.assertEqual("$1:$1", worksheet.print_title_rows)
        self.assertEqual("landscape", worksheet.page_setup.orientation)
        instruction_text = " ".join(str(cell.value or "") for row in instructions for cell in row)
        self.assertIn("前导零", instruction_text)
        self.assertIn("设备编码", instruction_text)
        self.assertIn("兼容旧表", instruction_text)

    def test_windows_build_copies_template_beside_executable(self) -> None:
        script = (self.project_root / "scripts" / "build_windows.ps1").read_text("utf-8")

        self.assertIn("站位表导入模板.xlsx", script)
        self.assertIn("Copy-Item", script)


if __name__ == "__main__":
    unittest.main()
