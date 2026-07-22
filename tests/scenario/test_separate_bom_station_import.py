import os
import sqlite3
import unittest
from pathlib import Path
from tempfile import TemporaryDirectory

os.environ.setdefault("QT_QPA_PLATFORM", "offscreen")

from openpyxl import Workbook
from PySide6.QtWidgets import QApplication

from smt_guard.bom import BomDocument, Material, Product
from smt_guard.configuration import ImportValidationError
from smt_guard.importing import ConfigurationImportService, ImportResult
from smt_guard.scan import ProductConfiguration
from smt_guard.sqlite import (
    SqliteDatabase,
    SqliteMasterDataRepository,
    SqliteProductConfigurationRepository,
)
from smt_guard.ui.importing import ConfigurationImportWidget
from smt_guard.xlsx_reader import OpenpyxlWorkbookReader


def sample_document() -> BomDocument:
    return BomDocument(
        product=Product("501000087", "BOM-1", "Control board", "控制板", "大板"),
        materials={
            "013000081": Material(
                "013000081", "端子线", "26AWG", "1", "电子物料/电子线"
            )
        },
    )


def sample_result() -> ImportResult:
    return ImportResult(
        sample_document(),
        ProductConfiguration(
            "501000087", "V1", {("SMT-01", "F-01"): "013000081"}
        ),
    )


class FakeSeparateImportWorkflow:
    def __init__(self) -> None:
        self.calls: list[tuple[object, ...]] = []

    def import_bom(self, bom_path: Path, *, version: str | None = None) -> BomDocument:
        del version
        self.calls.append(("bom", bom_path))
        return sample_document()

    def import_station_table(
        self,
        station_path: Path,
        *,
        version: str,
        station_sheet: str,
    ) -> ImportResult:
        self.calls.append(("stations", station_path, version, station_sheet))
        return sample_result()

    def import_configuration(
        self,
        station_path: Path,
        *,
        product_code: str,
        version: str,
        station_sheet: str,
    ) -> ImportResult:
        self.calls.append(
            ("configuration", station_path, product_code, version, station_sheet)
        )
        return ImportResult(None, sample_result().configuration)


class SeparateImportServiceTests(unittest.TestCase):
    def setUp(self) -> None:
        temporary = TemporaryDirectory()
        self.addCleanup(temporary.cleanup)
        self.directory = Path(temporary.name)
        self.connection = sqlite3.connect(":memory:")
        self.addCleanup(self.connection.close)
        SqliteDatabase(self.connection).initialize()
        self.master_data = SqliteMasterDataRepository(self.connection)
        self.master_data.add_device("SMT-01", "Machine 1", "Line A")
        self.master_data.add_station("SMT-01", "F-01")
        self.configurations = SqliteProductConfigurationRepository(self.connection)
        self.service = ConfigurationImportService(
            OpenpyxlWorkbookReader(), self.master_data, self.configurations
        )

    def write_bom(self) -> Path:
        path = self.directory / "bom.xlsx"
        workbook = Workbook()
        worksheet = workbook.active
        if worksheet is None:
            raise AssertionError("New workbook did not create a default worksheet")
        worksheet.title = "Worksheet"
        worksheet.append(
            [
                "BOM编号",
                "BOM名称",
                "深度",
                "单位用量",
                "商品编号",
                "商品名",
                "商品规格",
                "商品分类",
            ]
        )
        worksheet.append(
            ["BOM-1", "Control board", "0", "1", "501000087", "控制板", "大板", ""]
        )
        worksheet.append(
            ["", "", "1", "1", "013000081", "端子线", "26AWG", "电子物料/电子线"]
        )
        workbook.save(path)
        workbook.close()
        return path

    def write_invalid_bom(self) -> Path:
        path = self.directory / "invalid-bom.xlsx"
        workbook = Workbook()
        worksheet = workbook.active
        if worksheet is None:
            raise AssertionError("New workbook did not create a default worksheet")
        worksheet.title = "Worksheet"
        worksheet.append(["深度", "商品编号"])
        worksheet.append(["1", "013000081"])
        workbook.save(path)
        workbook.close()
        return path

    def write_stations(self) -> Path:
        path = self.directory / "stations.xlsx"
        workbook = Workbook()
        worksheet = workbook.active
        if worksheet is None:
            raise AssertionError("New workbook did not create a default worksheet")
        worksheet.title = "Worksheet"
        worksheet.append(["设备编码", "站位编码", "物料编码"])
        worksheet.append(["SMT-01", "F-01", "013000081"])
        workbook.save(path)
        workbook.close()
        return path

    def test_imports_bom_without_creating_configuration(self) -> None:
        document = self.service.import_bom(self.write_bom())

        self.assertEqual("501000087", document.product.material_code)
        self.assertEqual(["013000081"], list(document.materials))
        self.assertEqual([], self.configurations.list_configurations())

    def test_imports_station_table_later_using_loaded_bom(self) -> None:
        self.service.import_bom(self.write_bom())

        result = self.service.import_station_table(
            self.write_stations(), version=" V1 ", station_sheet=" Worksheet "
        )

        stored = self.configurations.get("501000087", "V1")
        self.assertEqual(result.document, sample_document())
        self.assertEqual("013000081", stored.required_material("SMT-01", "F-01"))

    def test_rejects_station_table_before_bom(self) -> None:
        with self.assertRaises(ImportValidationError) as caught:
            self.service.import_station_table(
                self.write_stations(), version="V1", station_sheet="Worksheet"
            )

        self.assertIn("BOM", str(caught.exception))
        self.assertEqual([], self.configurations.list_configurations())

    def test_failed_new_bom_import_clears_previously_loaded_bom(self) -> None:
        self.service.import_bom(self.write_bom())

        with self.assertRaises(ValueError):
            self.service.import_bom(self.write_invalid_bom())

        with self.assertRaises(ImportValidationError) as caught:
            self.service.import_station_table(
                self.write_stations(), version="V1", station_sheet="Worksheet"
            )
        self.assertIn("BOM", str(caught.exception))


class SeparateImportWidgetTests(unittest.TestCase):
    app: QApplication

    @classmethod
    def setUpClass(cls) -> None:
        application = QApplication.instance()
        if application is None:
            cls.app = QApplication([])
        elif isinstance(application, QApplication):
            cls.app = application
        else:
            raise RuntimeError("A non-GUI Qt application already exists")

    def test_widget_uses_single_direct_configuration_action(self) -> None:
        workflow = FakeSeparateImportWorkflow()
        widget = ConfigurationImportWidget(workflow)
        self.addCleanup(widget.close)
        widget.product_code_input.setText(" 501000087 ")
        widget.station_path_input.setText(" C:/imports/stations.xlsx ")
        widget.version_input.setText(" V1 ")
        widget.station_import_button.click()

        self.assertEqual(
            (
                "configuration",
                Path("C:/imports/stations.xlsx"),
                "501000087",
                "V1",
                "",
            ),
            workflow.calls[-1],
        )
        self.assertEqual(1, widget.assignment_table.rowCount())
        self.assertEqual("success", widget.status_label.property("feedbackState"))

    def test_widget_has_no_bom_import_controls(self) -> None:
        workflow = FakeSeparateImportWorkflow()
        widget = ConfigurationImportWidget(workflow)
        self.addCleanup(widget.close)
        self.assertFalse(hasattr(widget, "bom_path_input"))
        self.assertFalse(hasattr(widget, "bom_import_button"))
        self.assertFalse(hasattr(widget, "station_sheet_input"))


if __name__ == "__main__":
    unittest.main()
