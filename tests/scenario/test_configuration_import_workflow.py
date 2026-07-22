import sqlite3
import unittest
from pathlib import Path
from tempfile import TemporaryDirectory

from openpyxl import Workbook

from smt_guard.bom import BomStatus
from smt_guard.configuration import ConfigurationStatus, ImportValidationError
from smt_guard.importing import ConfigurationImportService
from smt_guard.sqlite import (
    SqliteBomRepository,
    SqliteDatabase,
    SqliteMasterDataRepository,
    SqliteProductConfigurationRepository,
    UnknownConfigurationError,
)
from smt_guard.xlsx_reader import OpenpyxlWorkbookReader, WorkbookReadError


class ConfigurationImportWorkflowTests(unittest.TestCase):
    def setUp(self) -> None:
        temporary = TemporaryDirectory()
        self.addCleanup(temporary.cleanup)
        self.directory = Path(temporary.name)
        self.connection = sqlite3.connect(":memory:")
        self.addCleanup(self.connection.close)
        SqliteDatabase(self.connection).initialize()
        self.master_data = SqliteMasterDataRepository(self.connection)
        self.master_data.add_device("SMT-01", "Machine 1", "Line A")
        self.master_data.add_station("SMT-01", "F-01")
        self.configurations = SqliteProductConfigurationRepository(self.connection)
        self.service = ConfigurationImportService(
            OpenpyxlWorkbookReader(), self.master_data, self.configurations
        )

    def write_bom(self) -> Path:
        path = self.directory / "bom.xlsx"
        workbook = Workbook()
        worksheet = workbook.active
        if worksheet is None:
            raise AssertionError("New workbook did not create a default worksheet")
        worksheet.title = "Worksheet"
        worksheet.append(
            ["BOM编号", "BOM名称", "深度", "单位用量", "商品编号", "商品名", "商品规格", "商品分类"]
        )
        worksheet.append(
            ["BOM-1", "Control board", "0", "1", "501000087", "控制板", "大板", ""]
        )
        worksheet.append(
            ["", "", "1", "1", "013000081", "端子线", "26AWG", "电子物料/电子线"]
        )
        workbook.save(path)
        workbook.close()
        return path

    def write_invalid_bom(self) -> Path:
        path = self.directory / "invalid-bom.xlsx"
        workbook = Workbook()
        worksheet = workbook.active
        if worksheet is None:
            raise AssertionError("New workbook did not create a default worksheet")
        worksheet.title = "Worksheet"
        worksheet.append(["深度", "商品编号"])
        worksheet.append(["1", "013000081"])
        workbook.save(path)
        workbook.close()
        return path

    def write_stations(self, material_code: str = "013000081") -> Path:
        path = self.directory / "stations.xlsx"
        workbook = Workbook()
        worksheet = workbook.active
        if worksheet is None:
            raise AssertionError("New workbook did not create a default worksheet")
        worksheet.title = "Stations"
        worksheet.append(["设备编码", "站位编码", "物料编码"])
        worksheet.append(["SMT-01", "F-01", material_code])
        workbook.save(path)
        workbook.close()
        return path

    def test_imports_real_workbooks_and_persists_exact_assignment(self) -> None:
        result = self.service.import_files(
            self.write_bom(),
            self.write_stations(),
            version=" V1 ",
            station_sheet="Stations",
        )

        stored = self.configurations.get("501000087", "V1")
        assert result.document is not None
        self.assertEqual("501000087", result.document.product.material_code)
        self.assertEqual(1, len(result.configuration.assignments))
        self.assertEqual("013000081", stored.required_material("SMT-01", "F-01"))

    def test_imports_two_column_configuration_without_bom(self) -> None:
        path = self.directory / "direct-configuration.xlsx"
        workbook = Workbook()
        sheet = workbook.active
        assert sheet is not None
        sheet.title = "Stations"
        sheet.append(["站位编码", "物料编码"])
        sheet.append(["F-01", "MATERIAL-NOT-IN-BOM"])
        workbook.save(path)
        workbook.close()

        result = self.service.import_configuration(
            path,
            product_code="501000087",
            version="V-DIRECT",
            station_sheet="Stations",
        )

        self.assertIsNone(result.document)
        self.assertEqual(
            "MATERIAL-NOT-IN-BOM",
            result.configuration.required_material("SMT-01", "F-01"),
        )
        self.assertEqual(
            ConfigurationStatus.ACTIVE,
            self.configurations.get_record("501000087", "V-DIRECT").status,
        )

    def test_full_import_activates_the_linked_bom_after_validation(self) -> None:
        boms = SqliteBomRepository(self.connection)
        service = ConfigurationImportService(
            OpenpyxlWorkbookReader(),
            self.master_data,
            self.configurations,
            boms,
            operator="OP-UI",
        )

        result = service.import_files(
            self.write_bom(),
            self.write_stations(),
            version="V1",
            station_sheet="Stations",
        )

        versions = boms.list_versions("501000087")
        self.assertEqual(1, len(versions))
        self.assertEqual(BomStatus.ACTIVE, versions[0].status)
        self.assertEqual(versions[0].id, result.configuration.bom_version_id)

    def test_legacy_three_column_import_no_longer_rejects_material_outside_bom(self) -> None:
        result = self.service.import_files(
            self.write_bom(),
            self.write_stations("999999999"),
            version="V1",
            station_sheet="Stations",
        )

        self.assertEqual(
            "999999999",
            result.configuration.required_material("SMT-01", "F-01"),
        )

    def test_reports_missing_station_worksheet_without_saving(self) -> None:
        with self.assertRaises(WorkbookReadError) as caught:
            self.service.import_files(
                self.write_bom(),
                self.write_stations(),
                version="V1",
                station_sheet="Missing",
            )

        self.assertIn("Missing", str(caught.exception))
        self.assertIn("Stations", str(caught.exception))
        with self.assertRaises(UnknownConfigurationError):
            self.configurations.get("501000087", "V1")

    def test_failed_new_bom_import_clears_previously_loaded_bom(self) -> None:
        self.service.import_bom(self.write_bom())

        with self.assertRaises(ValueError):
            self.service.import_bom(self.write_invalid_bom())

        with self.assertRaises(ImportValidationError) as caught:
            self.service.import_station_table(
                self.write_stations(), version="V1", station_sheet="Stations"
            )
        self.assertIn("BOM", str(caught.exception))


if __name__ == "__main__":
    unittest.main()
